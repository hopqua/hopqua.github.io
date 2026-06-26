#!/usr/bin/env python3
"""Cập nhật cột Mô tả sản phẩm trong Shopee_mass_upload_71sp.xlsx từ products.js."""
from __future__ import annotations

import argparse
import re
import shutil
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
PRODUCTS_JS = ROOT / "js" / "products.js"
DEFAULT_IN = Path("/home/vananh/.local/share/Trash/files/Shopee_mass_upload_71sp.xlsx")
DEFAULT_OUT = Path("/home/vananh/Downloads/Shopee/Shopee_mass_upload_71sp.xlsx")

SHEET = "Bản đăng tải"
DATA_START_ROW = 7
COL_DESC = 3  # C — Mô tả sản phẩm
COL_SKU = 4   # D — SKU sản phẩm (= product id)

FOOTER = (
    "Phụ kiện hộp / khay đựng bánh Trung Thu, hàng có sẵn.\n"
    "Inbox shop để được giá tốt theo số lượng."
)


def parse_products_js() -> dict[str, dict]:
    text = PRODUCTS_JS.read_text(encoding="utf-8")
    products: dict[str, dict] = {}
    for m in re.finditer(
        r"id:\s*'([^']+)',\s*name:\s*'([^']*)',[\s\S]*?"
        r"price:\s*'([^']*)',\s*description:\s*'((?:\\'|[^'])*)'",
        text,
    ):
        pid, _name, price, desc = m.group(1), m.group(2), m.group(3), m.group(4)
        desc = desc.replace("\\'", "'").replace("\\n", "\n")
        intro = desc
        spec = ""
        if "\n\n•" in desc:
            intro, spec = desc.split("\n\n", 1)
        elif "\n\n【" in desc:
            intro, spec = desc.split("\n\n", 1)
        products[pid] = {"price": price, "intro": intro.strip(), "spec": spec.strip()}
    return products


def build_shopee_description(intro: str, spec: str, price: str, old_desc: str | None) -> str:
    intro = intro.strip()
    if intro and not intro.endswith("."):
        intro += "."

    if not price and old_desc:
        m = re.search(r"Mức giá tham khảo:\s*([^\n]+)", old_desc)
        if m:
            price = m.group(1).strip().rstrip(".")

    parts: list[str] = []
    if intro:
        parts.append(intro)
    if spec:
        parts.extend(["", spec])
    parts.extend(["", f"Mức giá tham khảo: {price}.", FOOTER])
    desc = "\n".join(parts)

    if len(desc) < 100:
        desc += (
            "\nXem thêm ảnh chi tiết tại shop. "
            "Zalo 0965671689 — tư vấn chọn mẫu và báo giá sỉ theo số lượng."
        )
    return desc[:3000]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, default=DEFAULT_IN)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUT)
    args = parser.parse_args()

    if not args.input.is_file():
        raise SystemExit(f"Không thấy file: {args.input}")

    catalog = parse_products_js()
    args.output.parent.mkdir(parents=True, exist_ok=True)
    if args.input.resolve() != args.output.resolve():
        shutil.copy2(args.input, args.output)

    wb = openpyxl.load_workbook(args.output)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"Không có sheet '{SHEET}'")
    ws = wb[SHEET]

    updated = 0
    missing = []
    for row in range(DATA_START_ROW, ws.max_row + 1):
        sku = ws.cell(row=row, column=COL_SKU).value
        if not sku:
            continue
        sku = str(sku).strip()
        old_desc = ws.cell(row=row, column=COL_DESC).value
        old_desc = str(old_desc) if old_desc else None

        prod = catalog.get(sku)
        if not prod:
            missing.append(sku)
            continue

        new_desc = build_shopee_description(
            prod["intro"], prod["spec"], prod["price"], old_desc
        )
        ws.cell(row=row, column=COL_DESC).value = new_desc
        updated += 1

    wb.save(args.output)
    print(f"Đã cập nhật {updated}/71 mô tả → {args.output}")
    if missing:
        print(f"Không khớp products.js ({len(missing)}):", ", ".join(missing[:10]), "...")


if __name__ == "__main__":
    main()
