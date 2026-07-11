#!/usr/bin/env python3
"""Tạo Excel Shopee sẵn đăng + báo cáo HTML giá (lẻ 1–10 × 1,30 phí sàn × 1,04 hoàn hàng)."""
from __future__ import annotations

import argparse
import csv
import html
import json
import math
import re
import shutil
import sys
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path

import openpyxl

SCRIPTS = Path(__file__).resolve().parent
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))
from product_urls import product_page_url

ROOT = Path(__file__).resolve().parent.parent
PRODUCTS_JS = ROOT / "js" / "products.js"
MANIFEST_JS = ROOT / "js" / "product-images-manifest.js"
def _hop_qua_root() -> Path:
    here = Path(__file__).resolve()
    for parent in here.parents:
        candidate = parent / "Shopee_mass_upload_71sp.xlsx"
        if candidate.is_file():
            return parent
        candidate = parent / "hop-qua" / "Shopee_mass_upload_71sp.xlsx"
        if candidate.is_file():
            return parent / "hop-qua"
    return Path("/home/vananh/huong-dan/du-an/vo-anh/hop-qua")


HOP_QUA = _hop_qua_root()
DEFAULT_TEMPLATE = HOP_QUA / "Shopee_mass_upload_71sp.xlsx"
DEFAULT_OUT_DIR = HOP_QUA / "shopee-upload-2026"

SITE = "https://hopqua.io.vn"
SHOP_ID = 1307955653
SHEET = "Bản đăng tải"
DATA_START_ROW = 7

COL_CAT = 1
COL_NAME = 2
COL_DESC = 3
COL_SKU = 4
COL_VAR_CODE = 5
COL_VAR_GROUP = 6
COL_VAR_OPT = 7
COL_VAR_IMG = 8
COL_PRICE = 11
COL_STOCK = 12
COL_VAR_SKU = 13
COL_COVER = 17
COL_IMG_START = 18
# Kênh vận chuyển Shopee (sheet «Bản đăng tải»)
COL_SHIP_FAST = 30       # Nhanh
COL_SHIP_BULKY = 31      # Hàng Cồng Kềnh
COL_SHIP_VIETTEL = 32    # Tủ nhận hàng - Viettel Smartbox
COL_SHIP_SPX = 33        # Tủ nhận hàng - SPX
COL_SHIP_PUDO = 34       # Điểm nhận hàng
SHIPPING_CHANNEL_COLS = (
    COL_SHIP_FAST,
    COL_SHIP_BULKY,
    COL_SHIP_VIETTEL,
    COL_SHIP_SPX,
    COL_SHIP_PUDO,
)

FEE_PLATFORM = 0.30
FEE_RETURN = 0.04
ROUND_STEP = 500

FOOTER = (
    "Phụ kiện hộp / khay đựng bánh Trung Thu, hàng có sẵn.\n"
    "Chat với shop trên Shopee để được tư vấn mẫu và ưu đãi theo số lượng.\n"
    "Ảnh chi tiết đầy đủ trong gallery sản phẩm."
)

# Dòng / URL dẫn giao dịch ngoài Shopee (vi phạm chính sách listing)
_OFFPLATFORM_RE = re.compile(
    r"https?://|hopqua\.github|zalo|facebook\.com|messenger|"
    r"0965671689|inbox shop|báo giá sỉ|liên hệ ngoài",
    re.I,
)


@dataclass
class Product:
    id: str
    name: str
    folder: str
    price_label: str
    description: str
    images: list[str] = field(default_factory=list)

    @property
    def intro(self) -> str:
        desc = self.description
        if "\n\n•" in desc:
            return desc.split("\n\n", 1)[0].strip()
        if "\n\n【" in desc:
            return desc.split("\n\n", 1)[0].strip()
        return desc.split("\n")[0].strip()

    @property
    def specs(self) -> str:
        desc = self.description
        if "\n\n•" in desc:
            return desc.split("\n\n", 1)[1].strip()
        if "\n\n【" in desc:
            return desc.split("\n\n", 1)[1].strip()
        return ""


def parse_products_js() -> dict[str, Product]:
    text = PRODUCTS_JS.read_text(encoding="utf-8")
    products: dict[str, Product] = {}
    for block in re.split(r"\n\s*\{", text)[1:]:
        pid = re.search(r"id:\s*'([^']+)'", block)
        if not pid:
            continue
        name = re.search(r"name:\s*'((?:\\'|[^'])*)'", block)
        folder = re.search(r"folder:\s*'([^']*)'", block)
        price = re.search(r"price:\s*'([^']*)'", block)
        desc = re.search(r"description:\s*'((?:\\'|[^'])*)'", block)
        products[pid.group(1)] = Product(
            id=pid.group(1),
            name=(name.group(1) if name else "").replace("\\'", "'"),
            folder=(folder.group(1) if folder else ""),
            price_label=(price.group(1) if price else ""),
            description=(desc.group(1) if desc else "").replace("\\'", "'").replace("\\n", "\n"),
        )
    return products


def parse_manifest() -> dict[str, list[str]]:
    text = MANIFEST_JS.read_text(encoding="utf-8")
    out: dict[str, list[str]] = {}
    for pid, body in re.findall(r"'([^']+)':\s*\[([\s\S]*?)\]", text):
        imgs = re.findall(r'"([^"]+)"', body)
        if imgs:
            out[pid] = imgs
    return out


def collect_vnd_amounts(text: str) -> list[int]:
    """Trích mọi mức giá (đồng) từ chuỗi — dùng max làm giá lẻ bán."""
    if not text:
        return []
    amounts: list[int] = []
    for m in re.finditer(r"([\d.]+)\s*đ", text, re.I):
        amounts.append(int(m.group(1).replace(".", "")))
    for m in re.finditer(r"(\d{1,3})\s*[-–]\s*(\d{1,3})\s*k\b", text, re.I):
        amounts.extend([int(m.group(1)) * 1000, int(m.group(2)) * 1000])
    for m in re.finditer(r"(\d+)\s*k\b", text, re.I):
        amounts.append(int(m.group(1)) * 1000)
    for m in re.finditer(r"(\d+)\s*d\s*[-–]\s*(\d+)\s*d\b", text, re.I):
        amounts.extend([int(m.group(1)), int(m.group(2))])
    return amounts


def parse_direct_retail_vnd(product: Product) -> int | None:
    """Giá lẻ bán = mức cao nhất trong khoảng (vd. 29–35k → 35.000đ)."""
    amounts: list[int] = []

    m = re.search(r"Giá lẻ \(1[–-]10 cái\):\s*([\d.]+)\s*đ", product.description, re.I)
    if m:
        amounts.append(int(m.group(1).replace(".", "")))

    amounts.extend(collect_vnd_amounts(product.price_label))
    amounts.extend(collect_vnd_amounts(product.id))
    amounts.extend(collect_vnd_amounts(product.name))

    if amounts:
        return max(amounts)
    return None


def round_up_step(value: float, step: int = ROUND_STEP) -> int:
    """Làm tròn lên bội step (vd. 25.350 → 25.500, 29.850 → 30.000)."""
    return int(math.ceil(value / step) * step)


def calc_shopee_price(direct: int) -> tuple[int, int, int, int]:
    fee30 = round(direct * FEE_PLATFORM)
    fee04 = round(direct * FEE_RETURN)
    raw = direct * (1 + FEE_PLATFORM) * (1 + FEE_RETURN)
    final = round_up_step(raw)
    return final, fee30, fee04, final - direct


def resolve_shopee_listing_price(direct: int | None) -> tuple[int | None, int, int, int]:
    """Trả về (giá Shopee, fee30, fee04, chênh so với giá lẻ)."""
    if direct is None:
        return None, 0, 0, 0
    return calc_shopee_price(direct)


def load_shopee_links() -> dict[str, str]:
    """SKU → URL từ js/shopee-links.js (website)."""
    path = ROOT / "js" / "shopee-links.js"
    if not path.is_file():
        return {}
    return dict(re.findall(r"'([^']+)':\s*'(https://shopee\.vn/product/\d+/\d+)'", path.read_text(encoding="utf-8")))


def _is_shopee_item_id(val: object) -> bool:
    return bool(re.fullmatch(r"\d{8,12}", str(val or "").strip()))


def read_seller_export_links(path: Path) -> tuple[dict[str, str], set[str]]:
    """Đọc mass_update_basic_info — SKU→link + tập item ID (kể cả dòng SKU trống)."""
    try:
        import pandas as pd
    except ImportError:
        return {}, set()
    if not path.is_file():
        return {}, set()
    df = pd.read_excel(path, engine="calamine", header=None)
    links: dict[str, str] = {}
    item_ids: set[str] = set()
    for i in range(len(df)):
        pid = df.iloc[i, 0]
        if not _is_shopee_item_id(pid):
            continue
        iid = str(int(float(pid)))
        item_ids.add(iid)
        url = f"https://shopee.vn/product/{SHOP_ID}/{iid}"
        sku = str(df.iloc[i, 1] or "").strip()
        if not sku or sku in ("nan", "SKU Sản phẩm", "et_title_parent_sku"):
            continue
        if re.fullmatch(r"[a-f0-9]{32}", sku):
            continue
        links[sku] = url
    return links, item_ids


def shopee_item_id(url: str) -> str:
    return url.rstrip("/").split("/")[-1] if url else ""


def is_on_shopee(
    sku: str,
    export_links: dict[str, str],
    export_item_ids: set[str],
    all_links: dict[str, str],
) -> bool:
    if sku in export_links:
        return True
    if export_item_ids:
        iid = shopee_item_id(all_links.get(sku, ""))
        return bool(iid and iid in export_item_ids)
    return sku in all_links


STOCK_STATUS_FILE = "stock-status.json"
PRICE_OVERRIDES_FILE = "gia-le-cap-nhat.json"


def load_price_overrides(out_dir: Path) -> dict[str, int]:
    """SKU → giá lẻ (đồng) từ file anh sửa trên báo cáo HTML."""
    path = out_dir / PRICE_OVERRIDES_FILE
    if not path.is_file():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}
    raw = data.get("prices") if isinstance(data, dict) else None
    if not isinstance(raw, dict):
        raw = data if isinstance(data, dict) else {}
    out: dict[str, int] = {}
    for key, val in raw.items():
        if key.startswith("_"):
            continue
        try:
            n = int(val)
            if n > 0:
                out[str(key)] = n
        except (TypeError, ValueError):
            continue
    return out


@dataclass
class GiaLeCsvRow:
    sku: str
    ten: str
    gia_le: int
    gia_shopee: int | None = None


def load_gia_le_csv(path: Path) -> list[GiaLeCsvRow]:
    """Đọc gia-le-cap-nhat.csv — thứ tự dòng = thứ tự báo cáo."""
    if not path.is_file():
        return []
    out: list[GiaLeCsvRow] = []
    with path.open(encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            sku = str(row.get("sku") or "").strip()
            if not sku:
                continue
            try:
                gia_le = int(float(row.get("gia_le_vnd") or 0))
            except (TypeError, ValueError):
                continue
            gia_shopee: int | None = None
            raw_sp = row.get("gia_shopee_vnd")
            if raw_sp not in (None, ""):
                try:
                    gia_shopee = int(float(raw_sp))
                except (TypeError, ValueError):
                    gia_shopee = None
            out.append(
                GiaLeCsvRow(
                    sku=sku,
                    ten=str(row.get("ten") or sku).strip(),
                    gia_le=gia_le,
                    gia_shopee=gia_shopee,
                )
            )
    return out


def load_shopee_prices_json(out_dir: Path) -> dict[str, int]:
    """SKU → giá Shopee từ gia-le-cap-nhat.json (nếu có)."""
    path = out_dir / PRICE_OVERRIDES_FILE
    if not path.is_file():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}
    raw = data.get("shopee_prices") if isinstance(data, dict) else None
    if not isinstance(raw, dict):
        return {}
    out: dict[str, int] = {}
    for key, val in raw.items():
        if key.startswith("_"):
            continue
        try:
            n = int(val)
            if n > 0:
                out[str(key)] = n
        except (TypeError, ValueError):
            continue
    return out


def filter_reports_by_gia_le(
    reports: list[RowReport],
    gia_rows: list[GiaLeCsvRow],
    shopee_prices: dict[str, int],
) -> list[RowReport]:
    """Chỉ giữ SKU có trong CSV; tên + giá lẻ/Shopee theo file anh sửa."""
    by_sku = {r.sku: r for r in reports}
    out: list[RowReport] = []
    for i, g in enumerate(gia_rows, start=1):
        base = by_sku.get(g.sku)
        if not base:
            continue
        direct = g.gia_le
        fee30 = round(direct * FEE_PLATFORM) if direct else 0
        fee04 = round(direct * FEE_RETURN) if direct else 0
        shopee = shopee_prices.get(g.sku) or g.gia_shopee
        if shopee is None and direct:
            shopee, fee30, fee04, _ = resolve_shopee_listing_price(direct)
        delta = (
            (shopee - base.old_price)
            if shopee is not None and base.old_price
            else base.delta
        )
        out.append(
            RowReport(
                row=i,
                sku=g.sku,
                name=g.ten,
                direct=direct,
                fee30=fee30,
                fee04=fee04,
                shopee_price=shopee,
                old_price=base.old_price,
                delta=delta,
                has_product=base.has_product,
                has_images=base.has_images,
                web_url=base.web_url,
                on_shopee=base.on_shopee,
                shopee_url=base.shopee_url,
                thumb_url=base.thumb_url,
                image_url=base.image_url,
                in_stock=base.in_stock,
                stock_label=base.stock_label,
                note=base.note,
            )
        )
    return out


def write_gia_le_csv(rows: list[RowReport], out_path: Path) -> None:
    """Mẫu CSV sửa giá bằng Excel — chạy apply-gia-le-cap-nhat.py sau khi sửa."""
    with out_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["sku", "ten", "gia_le_vnd", "gia_shopee_vnd", "ghi_chu"])
        for r in rows:
            w.writerow([r.sku, r.name, r.direct or "", r.shopee_price or "", ""])


def load_stock_status(out_dir: Path) -> dict[str, bool]:
    """SKU → còn hàng (true) / hết hàng (false). Sửa file này rồi chạy lại script."""
    path = out_dir / STOCK_STATUS_FILE
    if not path.is_file():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}
    if not isinstance(data, dict):
        return {}
    out: dict[str, bool] = {}
    for key, val in data.items():
        if key.startswith("_"):
            continue
        if isinstance(val, bool):
            out[key] = val
        elif isinstance(val, dict) and "inStock" in val:
            out[key] = bool(val["inStock"])
    return out


def resolve_stock_status(sku: str, stock_map: dict[str, bool]) -> tuple[bool, str]:
    in_stock = stock_map.get(sku, True)
    return in_stock, "Còn hàng" if in_stock else "Hết hàng"


def find_seller_export_file(out_dir: Path) -> Path | None:
    """Tìm file export basic info Seller Centre để đối chiếu đã đăng."""
    names = (
        "seller-doi-chieu-basic-info.xlsx",
        "mass_update_basic_info_SOURCE.xlsx",
    )
    for name in names:
        p = out_dir / name
        if p.is_file():
            return p
    for base in (out_dir, Path.home() / "Downloads"):
        if not base.is_dir():
            continue
        matches = sorted(
            base.glob("mass_update_basic_info_*.xlsx"),
            key=lambda x: x.stat().st_mtime,
            reverse=True,
        )
        if matches:
            return matches[0]
    return None


def load_posted_from_seller_export(out_dir: Path) -> tuple[dict[str, str], set[str], str]:
    """Trả về (SKU→URL, item ID trên shop, mô tả nguồn) từ export Seller Centre."""
    path = find_seller_export_file(out_dir)
    if not path:
        return {}, set(), ""
    links, item_ids = read_seller_export_links(path)
    if not item_ids:
        return {}, set(), ""
    note = (
        f"<strong>Đối chiếu «đã đăng Shopee»</strong> theo export Seller Centre: "
        f"<code>{path.name}</code> ({len(item_ids)} SP trên shop"
        f"{f', {len(links)} có SKU khớp' if len(links) < len(item_ids) else ''})."
    )
    return links, item_ids, note


def fmt_vnd(n: int | None) -> str:
    if n is None:
        return "—"
    return f"{n:,}".replace(",", ".") + "đ"


def abs_url(path: str) -> str:
    if path.startswith("http"):
        return path
    return f"{SITE}/{path.lstrip('/')}"


def thumb_url_for_image(path: str) -> str:
    """URL thumbnail (-thumb.jpg) nếu có, không thì ảnh gốc."""
    if not path:
        return ""
    if re.search(r"-thumb\.jpe?g$", path, re.I):
        return abs_url(path)
    if re.search(r"\.jpe?g$", path, re.I):
        return abs_url(re.sub(r"(\.jpe?g)$", r"-thumb\1", path, flags=re.I))
    return abs_url(path)


def product_thumb_url(prod: Product | None) -> str:
    if not prod or not prod.images:
        return ""
    return thumb_url_for_image(prod.images[0])


def sanitize_shopee_text(text: str) -> str:
    """Bỏ URL và nội dung gợi ý giao dịch ngoài Shopee."""
    if not text:
        return ""
    kept: list[str] = []
    for line in text.splitlines():
        if _OFFPLATFORM_RE.search(line):
            continue
        if re.search(r"Giá lẻ|Giá tham khảo|SL 11|SL 100|Trên 1\.000", line, re.I):
            continue
        kept.append(line)
    return "\n".join(kept).strip()


def enable_all_shipping_channels(ws, row: int) -> None:
    for col in SHIPPING_CHANNEL_COLS:
        ws.cell(row=row, column=col).value = "Bật"


def build_shopee_title(old_name: str, product: Product | None) -> str:
    name = (product.name if product else old_name).strip()
    if product and product.folder.startswith("cap-nhat-2026/"):
        prefix = "[Mẫu 2026]"
        if not name.lower().startswith("[mẫu"):
            short = name[:90]
            return f"{prefix} Vỏ hộp trung thu {short}"[:120]
    return (old_name or name)[:120]


def build_shopee_description(product: Product, shopee_price: int) -> str:
    intro = sanitize_shopee_text(product.intro)
    if intro and not intro.endswith("."):
        intro += "."
    specs_clean = sanitize_shopee_text(product.specs)

    parts: list[str] = []
    if intro:
        parts.append(intro)
    if specs_clean:
        parts.extend(["", specs_clean])
    parts.extend(
        [
            "",
            f"Giá Shopee (lẻ 1–10 cái): {fmt_vnd(shopee_price)}.",
            FOOTER,
        ]
    )
    desc = "\n".join(parts)
    if len(desc) < 100:
        desc += "\nChat với shop trên Shopee để được tư vấn chọn mẫu phù hợp."
    return desc[:3000]


def enrich_images(products: dict[str, Product], manifest: dict[str, list[str]]) -> None:
    js = PRODUCTS_JS.read_text(encoding="utf-8")
    for pid, p in products.items():
        imgs = manifest.get(pid, [])
        if not imgs:
            m = re.search(
                rf"id:\s*'{re.escape(pid)}'[\s\S]*?thumbnail:\s*'([^']+)'",
                js,
            )
            if m:
                imgs = [m.group(1)]
        p.images = imgs


@dataclass
class RowReport:
    row: int
    sku: str
    name: str
    direct: int | None
    fee30: int
    fee04: int
    shopee_price: int | None
    old_price: int | None
    delta: int | None
    has_product: bool
    has_images: bool
    web_url: str
    on_shopee: bool
    shopee_url: str
    thumb_url: str = ""
    image_url: str = ""
    in_stock: bool = True
    stock_label: str = ""
    note: str = ""


def write_html_report(
    rows: list[RowReport],
    out_path: Path,
    formula_note: str,
    *,
    excel_name: str = "Shopee_mass_upload_71sp_READY.xlsx",
    page_title: str = "Báo cáo giá listing Shopee — mùa Trung Thu 2026",
    skip_note: str = "",
    doi_chieu_note: str = "",
    stock_note: str = "",
) -> None:
    updated = sum(1 for r in rows if r.has_product and r.direct)
    missing = [r for r in rows if not r.has_product]
    no_price = [r for r in rows if r.has_product and not r.direct]
    posted = sum(1 for r in rows if r.on_shopee)
    not_posted = len(rows) - posted
    in_stock_count = sum(1 for r in rows if r.in_stock)
    out_stock_count = len(rows) - in_stock_count
    default_stock = {r.sku: r.in_stock for r in rows}
    default_stock_json = json.dumps(default_stock, ensure_ascii=False)
    default_prices = {r.sku: r.direct for r in rows if r.direct}
    default_prices_json = json.dumps(default_prices, ensure_ascii=False)
    fee_platform = FEE_PLATFORM
    fee_return = FEE_RETURN
    round_step = ROUND_STEP

    def shopee_status_cell(r: RowReport) -> str:
        if r.on_shopee:
            return f'<span class="badge badge-yes">Đã đăng</span> <a href="{r.shopee_url}" target="_blank" rel="noopener" class="shopee-mini">↗</a>'
        return '<span class="badge badge-no">Chưa đăng</span>'

    def tr(r: RowReport) -> str:
        cls = ""
        if not r.has_product:
            cls = "warn"
        elif r.delta and abs(r.delta) > 500:
            cls = "delta"
        if r.on_shopee:
            cls = (cls + " on-shopee").strip()
        if not r.in_stock:
            cls = (cls + " out-of-stock").strip()
        sku_attr = html.escape(r.sku, quote=True)
        checked = " checked" if r.in_stock else ""
        stock_cell = (
            f'<label class="stock-check" title="Tích ✓ = còn hàng">'
            f'<input type="checkbox" class="stock-cb" data-sku="{sku_attr}"{checked}>'
            f'<span class="stock-mark" aria-hidden="true">✓</span></label>'
        )
        if r.thumb_url:
            fallback = r.image_url or r.thumb_url
            img_html = (
                f'<a href="{r.web_url}" target="_blank" rel="noopener" title="{r.sku}">'
                f'<img src="{r.thumb_url}" alt="{r.name}" width="80" height="80" loading="lazy" '
                f'onerror="this.onerror=null;this.src=\'{fallback}\';"></a>'
            )
        else:
            img_html = '<span class="no-img">—</span>'
        sku_hint = f'<code class="sku-mini" title="{r.sku}">{r.sku[:28]}{"…" if len(r.sku) > 28 else ""}</code>'

        direct_val = r.direct if r.direct else ""
        old_excel = r.old_price if r.old_price else ""
        name_attr = html.escape(r.name, quote=True)
        price_input = (
            f'<input type="number" class="price-input" data-sku="{sku_attr}" '
            f'value="{direct_val}" min="0" step="500" data-old-excel="{old_excel}" '
            f'aria-label="Giá lẻ {name_attr}" title="Giá lẻ 1–10 (đồng)">'
        )
        fee30_disp = fmt_vnd(r.fee30) if r.direct else "—"
        fee04_disp = fmt_vnd(r.fee04) if r.direct else "—"
        shopee_disp = fmt_vnd(r.shopee_price) if r.shopee_price else "—"
        delta_disp = fmt_vnd(r.delta) if r.delta is not None else "—"

        return f"""<tr class="{cls}" data-sku="{sku_attr}" data-name="{name_attr}">
          <td>{r.row}</td>
          <td class="img-cell">{img_html}{sku_hint}</td>
          <td class="stock-cell">{stock_cell}</td>
          <td><a href="{r.web_url}" target="_blank" rel="noopener">{r.name}</a></td>
          <td class="center">{shopee_status_cell(r)}</td>
          <td class="num price-cell">{price_input}</td>
          <td class="num fee30-cell">{fee30_disp}</td>
          <td class="num fee04-cell">{fee04_disp}</td>
          <td class="num"><strong class="shopee-price-cell">{shopee_disp}</strong></td>
          <td class="num excel-old-cell">{fmt_vnd(r.old_price) if r.old_price else '—'}</td>
          <td class="num delta-cell">{delta_disp}</td>
          <td><a href="{r.web_url}" target="_blank" rel="noopener">Xem SP</a>{(' · ' + r.note) if r.note else ''}</td>
        </tr>"""

    html_doc = f"""<!DOCTYPE html>
<html lang="vi">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{page_title} — Vân Thắng</title>
  <style>
    :root {{ --bg:#fff8f0; --text:#2d1810; --accent:#8b1528; --gold:#f0c14b; --border:#e8d5c0; }}
    body {{ font-family: system-ui, sans-serif; background: var(--bg); color: var(--text); margin: 0; padding: 24px; line-height: 1.5; }}
    h1 {{ color: var(--accent); font-size: 1.5rem; }}
    .meta {{ background: #fff; border: 1px solid var(--border); border-radius: 12px; padding: 16px 20px; margin-bottom: 20px; }}
    .meta strong {{ color: var(--accent); }}
    table {{ width: 100%; border-collapse: collapse; background: #fff; font-size: 0.88rem; box-shadow: 0 4px 16px rgba(45,24,16,.08); border-radius: 12px; overflow: hidden; }}
    th, td {{ border-bottom: 1px solid var(--border); padding: 8px 10px; text-align: left; vertical-align: top; }}
    th {{ background: linear-gradient(135deg, #9b1c31, #6d1222); color: #fff; position: sticky; top: 0; }}
    tr.on-shopee {{ background: #f0fdf4; }}
    tr.out-of-stock {{ background: #fef2f2; }}
    tr.out-of-stock.on-shopee {{ background: #fff1f2; }}
    tr.warn {{ background: #fff5f5; }}
    tr.delta {{ background: #fffbeb; }}
    .center {{ text-align: center; white-space: nowrap; }}
    .badge {{ display: inline-block; padding: 2px 8px; border-radius: 999px; font-size: 0.78rem; font-weight: 700; }}
    .badge-yes {{ background: #dcfce7; color: #166534; border: 1px solid #86efac; }}
    .badge-no {{ background: #f3f4f6; color: #6b7280; border: 1px solid #d1d5db; }}
    .stock-cell {{ width: 56px; text-align: center; vertical-align: middle; }}
    .stock-check {{ display: inline-flex; align-items: center; justify-content: center; width: 36px; height: 36px; cursor: pointer; position: relative; }}
    .stock-check input {{ position: absolute; opacity: 0; width: 100%; height: 100%; margin: 0; cursor: pointer; }}
    .stock-mark {{ display: flex; align-items: center; justify-content: center; width: 28px; height: 28px; border: 2px solid #d1d5db; border-radius: 6px; background: #fff; color: transparent; font-size: 1.1rem; font-weight: 700; transition: .15s; }}
    .stock-check input:checked + .stock-mark {{ background: #dcfce7; border-color: #22c55e; color: #166534; }}
    .stock-check input:focus-visible + .stock-mark {{ outline: 2px solid var(--accent); outline-offset: 2px; }}
    .stock-toolbar {{ display: flex; flex-wrap: wrap; align-items: center; gap: 12px; margin: 0 0 16px; padding: 12px 16px; background: #fff; border: 1px solid var(--border); border-radius: 12px; }}
    .btn-save {{ background: var(--accent); color: #fff; border: none; padding: 10px 18px; border-radius: 8px; font-weight: 700; font-size: 0.95rem; cursor: pointer; }}
    .btn-save:hover {{ filter: brightness(1.08); }}
    .save-status {{ color: #166534; font-size: 0.9rem; font-weight: 600; }}
    .save-status.err {{ color: #991b1b; }}
    .shopee-mini {{ font-size: 0.85rem; text-decoration: none; margin-left: 2px; }}
    .num {{ text-align: right; white-space: nowrap; }}
    code {{ font-size: 0.8rem; }}
    .stats {{ display: flex; flex-wrap: wrap; gap: 12px; margin: 12px 0; }}
    .stat {{ background: var(--gold); color: var(--text); padding: 8px 14px; border-radius: 999px; font-weight: 600; font-size: 0.9rem; }}
    a {{ color: var(--accent); }}
    .img-cell {{ width: 96px; text-align: center; }}
    .img-cell img {{ display: block; width: 80px; height: 80px; object-fit: cover; border-radius: 8px; border: 1px solid var(--border); margin: 0 auto 4px; background: #fff; }}
    .img-cell .no-img {{ display: block; width: 80px; height: 80px; line-height: 80px; margin: 0 auto 4px; background: #f3f4f6; border-radius: 8px; color: #9ca3af; }}
    .sku-mini {{ display: block; font-size: 0.68rem; color: #6b7280; word-break: break-all; line-height: 1.25; max-width: 92px; margin: 0 auto; }}
    .price-input {{ width: 6.5rem; text-align: right; font: inherit; font-weight: 600; padding: 6px 8px; border: 2px solid #f0c14b; border-radius: 8px; background: #fffbeb; }}
    .price-input:focus {{ outline: 2px solid var(--accent); border-color: var(--accent); background: #fff; }}
    .price-input:placeholder-shown {{ font-weight: 400; }}
    tr.price-changed .price-input {{ border-color: #f59e0b; background: #fef3c7; }}
    .btn-save-secondary {{ background: #fff; color: var(--accent); border: 2px solid var(--accent); }}
    .btn-save-secondary:hover {{ background: #fff5f5; }}
  </style>
</head>
<body>
  <h1>{page_title}</h1>
  <div class="meta">
    <p>Ngày tạo: <strong>{date.today().isoformat()}</strong></p>
    {f'<p>{doi_chieu_note}</p>' if doi_chieu_note else ''}
    {f'<p>{stock_note}</p>' if stock_note else ''}
    {f'<p>{skip_note}</p>' if skip_note else ''}
    <p>Công thức: <strong>{formula_note}</strong></p>
    <p>File Excel sẵn đăng: <strong>{excel_name}</strong> (sheet «Bản đăng tải»)</p>
    <p>Sau khi đăng Shopee mới, export lại <strong>Thông tin cơ bản</strong> từ Seller Centre → chạy
    <code>python3 scripts/build-shopee-upload-package.py --seller-export ~/Downloads/mass_update_basic_info_*.xlsx</code></p>
    <div class="stats">
      <span class="stat">{len(rows)} dòng Excel</span>
      <span class="stat">{updated} đã tính giá mới</span>
      <span class="stat">{len(missing)} SKU chưa có trong products.js</span>
      <span class="stat">{len(no_price)} thiếu giá lẻ 1–10</span>
      <span class="stat">{posted} đã đăng Shopee</span>
      <span class="stat">{not_posted} chưa đăng Shopee</span>
      <span class="stat" id="stat-in-stock">{in_stock_count} còn hàng</span>
      <span class="stat" id="stat-out-stock">{out_stock_count} hết hàng</span>
    </div>
    <p><strong>Giá lẻ:</strong> sửa ô vàng cột «Giá lẻ» → tự tính Shopee · bấm <strong>Lưu giá</strong> tải <code>gia-le-cap-nhat.json</code> + <code>gia-le-cap-nhat.csv</code> (hoặc sửa CSV rồi chạy <code>apply-gia-le-cap-nhat.py</code>).</p>
    <div class="stock-toolbar">
      <button type="button" class="btn-save btn-save-secondary" id="btn-save-prices">Lưu giá</button>
      <span id="price-save-status" class="save-status"></span>
      <span style="color:#6b7280;font-size:0.85rem">Giá lẻ (đồng) · Shopee = lẻ × 1,30 × 1,04, làm tròn lên 500đ</span>
    </div>
    <div class="stock-toolbar">
      <button type="button" class="btn-save" id="btn-save-stock">Lưu tồn kho</button>
      <span id="save-status" class="save-status"></span>
      <span style="color:#6b7280;font-size:0.85rem">Tích ✓ = còn hàng · bỏ tích = hết hàng (ẩn sau)</span>
    </div>
  </div>
  <table id="stock-table">
    <thead>
      <tr>
        <th>#</th><th>Ảnh / SKU</th><th title="Tích = còn hàng">✓</th><th>Tên (link web)</th><th>Shopee</th>
        <th>Giá lẻ (nhập)</th><th>+30% sàn</th><th>+4% hoàn</th>
        <th>Giá Shopee mới</th><th>Giá Excel cũ</th><th>Chênh</th><th>Web</th>
      </tr>
    </thead>
    <tbody>
      {''.join(tr(r) for r in rows)}
    </tbody>
  </table>
  <script>
    (function () {{
      const STORAGE_STOCK = 'hopqua-shopee-bao-cao-stock';
      const STORAGE_PRICES = 'hopqua-shopee-bao-cao-prices';
      const DEFAULT_STOCK = {default_stock_json};
      const DEFAULT_PRICES = {default_prices_json};
      const FEE_PLATFORM = {fee_platform};
      const FEE_RETURN = {fee_return};
      const ROUND_STEP = {round_step};
      const table = document.getElementById('stock-table');
      const statusEl = document.getElementById('save-status');
      const priceStatusEl = document.getElementById('price-save-status');
      const statIn = document.getElementById('stat-in-stock');
      const statOut = document.getElementById('stat-out-stock');

      function fmtVnd(n) {{
        if (n === null || n === undefined || n === '' || isNaN(n)) return '—';
        return Number(n).toLocaleString('vi-VN') + 'đ';
      }}

      function calcShopee(direct) {{
        const d = parseInt(direct, 10);
        if (!d || d <= 0) return null;
        const fee30 = Math.round(d * FEE_PLATFORM);
        const fee04 = Math.round(d * FEE_RETURN);
        const raw = d * (1 + FEE_PLATFORM) * (1 + FEE_RETURN);
        const shopee = Math.ceil(raw / ROUND_STEP) * ROUND_STEP;
        return {{ fee30, fee04, shopee }};
      }}

      function recalcRow(row) {{
        const input = row.querySelector('.price-input');
        if (!input) return;
        const direct = parseInt(input.value, 10);
        const calc = calcShopee(direct);
        const fee30El = row.querySelector('.fee30-cell');
        const fee04El = row.querySelector('.fee04-cell');
        const shopeeEl = row.querySelector('.shopee-price-cell');
        const deltaEl = row.querySelector('.delta-cell');
        const oldExcel = parseInt(input.dataset.oldExcel || '0', 10) || 0;
        if (!calc) {{
          if (fee30El) fee30El.textContent = '—';
          if (fee04El) fee04El.textContent = '—';
          if (shopeeEl) shopeeEl.textContent = '—';
          if (deltaEl) deltaEl.textContent = '—';
          row.classList.remove('delta', 'price-changed');
          return;
        }}
        if (fee30El) fee30El.textContent = fmtVnd(calc.fee30);
        if (fee04El) fee04El.textContent = fmtVnd(calc.fee04);
        if (shopeeEl) shopeeEl.textContent = fmtVnd(calc.shopee);
        const delta = oldExcel ? calc.shopee - oldExcel : null;
        if (deltaEl) deltaEl.textContent = delta !== null ? fmtVnd(delta) : '—';
        row.classList.toggle('delta', delta !== null && Math.abs(delta) > 500);
        const sku = input.dataset.sku;
        const orig = DEFAULT_PRICES[sku];
        row.classList.toggle('price-changed', orig && direct !== orig);
      }}

      function recalcAll() {{
        table.querySelectorAll('tbody tr').forEach(recalcRow);
      }}

      function readSavedPrices() {{
        try {{
          const raw = localStorage.getItem(STORAGE_PRICES);
          return raw ? JSON.parse(raw) : null;
        }} catch (_) {{
          return null;
        }}
      }}

      function collectPrices() {{
        const prices = {{}};
        const shopee_prices = {{}};
        const items = [];
        table.querySelectorAll('tbody tr').forEach(function (row) {{
          const input = row.querySelector('.price-input');
          if (!input) return;
          const sku = input.dataset.sku;
          const direct = parseInt(input.value, 10);
          if (!direct || direct <= 0) return;
          const calc = calcShopee(direct);
          prices[sku] = direct;
          if (calc) shopee_prices[sku] = calc.shopee;
          items.push({{
            sku: sku,
            ten: row.dataset.name || '',
            gia_le_vnd: direct,
            gia_shopee_vnd: calc ? calc.shopee : '',
          }});
        }});
        return {{ prices, shopee_prices, items }};
      }}

      function applyPrices(saved) {{
        if (!saved) return;
        table.querySelectorAll('.price-input').forEach(function (input) {{
          const sku = input.dataset.sku;
          if (Object.prototype.hasOwnProperty.call(saved, sku)) {{
            input.value = saved[sku];
          }}
        }});
        recalcAll();
      }}

      function downloadBlob(filename, content, mime) {{
        const blob = new Blob([content], {{ type: mime }});
        const a = document.createElement('a');
        a.href = URL.createObjectURL(blob);
        a.download = filename;
        a.click();
        URL.revokeObjectURL(a.href);
      }}

      function toCsv(items) {{
        const esc = function (s) {{
          const t = String(s ?? '');
          return /[",\\n]/.test(t) ? '"' + t.replace(/"/g, '""') + '"' : t;
        }};
        const lines = ['sku,ten,gia_le_vnd,gia_shopee_vnd,ghi_chu'];
        items.forEach(function (it) {{
          lines.push([it.sku, it.ten, it.gia_le_vnd, it.gia_shopee_vnd, ''].map(esc).join(','));
        }});
        return lines.join('\\n') + '\\n';
      }}

      table.addEventListener('input', function (e) {{
        if (!e.target.classList.contains('price-input')) return;
        recalcRow(e.target.closest('tr'));
        priceStatusEl.textContent = '';
        priceStatusEl.classList.remove('err');
      }});

      document.getElementById('btn-save-prices').addEventListener('click', function () {{
        const {{ prices, shopee_prices, items }} = collectPrices();
        const payload = {{
          updated: new Date().toISOString().slice(0, 10),
          note: 'Giá lẻ 1–10 (đồng). Chạy: python3 scripts/apply-gia-le-cap-nhat.py',
          prices: prices,
          shopee_prices: shopee_prices,
        }};
        try {{
          localStorage.setItem(STORAGE_PRICES, JSON.stringify(prices));
        }} catch (err) {{
          priceStatusEl.textContent = 'Không lưu trình duyệt: ' + err.message;
          priceStatusEl.classList.add('err');
          return;
        }}
        downloadBlob('gia-le-cap-nhat.json', JSON.stringify(payload, null, 2), 'application/json');
        setTimeout(function () {{
          downloadBlob('gia-le-cap-nhat.csv', toCsv(items), 'text/csv;charset=utf-8');
        }}, 200);
        priceStatusEl.textContent = 'Đã lưu · tải gia-le-cap-nhat.json + .csv';
        priceStatusEl.classList.remove('err');
      }});

      function readSaved() {{
        try {{
          const raw = localStorage.getItem(STORAGE_STOCK);
          return raw ? JSON.parse(raw) : null;
        }} catch (_) {{
          return null;
        }}
      }}

      function collectStock() {{
        const out = {{}};
        table.querySelectorAll('.stock-cb').forEach(function (cb) {{
          out[cb.dataset.sku] = cb.checked;
        }});
        return out;
      }}

      function applyRowStyle(row, inStock) {{
        row.classList.toggle('out-of-stock', !inStock);
      }}

      function updateStats() {{
        const stock = collectStock();
        const vals = Object.values(stock);
        const inCount = vals.filter(Boolean).length;
        statIn.textContent = inCount + ' còn hàng';
        statOut.textContent = (vals.length - inCount) + ' hết hàng';
      }}

      function applyStock(stock) {{
        table.querySelectorAll('tbody tr').forEach(function (row) {{
          const cb = row.querySelector('.stock-cb');
          if (!cb) return;
          const sku = cb.dataset.sku;
          const inStock = Object.prototype.hasOwnProperty.call(stock, sku)
            ? !!stock[sku]
            : !!DEFAULT_STOCK[sku];
          cb.checked = inStock;
          applyRowStyle(row, inStock);
        }});
        updateStats();
      }}

      table.addEventListener('change', function (e) {{
        if (!e.target.classList.contains('stock-cb')) return;
        const row = e.target.closest('tr');
        applyRowStyle(row, e.target.checked);
        updateStats();
        statusEl.textContent = '';
        statusEl.classList.remove('err');
      }});

      document.getElementById('btn-save-stock').addEventListener('click', function () {{
        const stock = collectStock();
        try {{
          localStorage.setItem(STORAGE_STOCK, JSON.stringify(stock));
        }} catch (err) {{
          statusEl.textContent = 'Không lưu được trên trình duyệt: ' + err.message;
          statusEl.classList.add('err');
          return;
        }}
        downloadBlob('stock-status.json', JSON.stringify(stock, null, 2), 'application/json');
        statusEl.textContent = 'Đã lưu · đã tải stock-status.json';
        statusEl.classList.remove('err');
      }});

      applyPrices(readSavedPrices() || DEFAULT_PRICES);
      recalcAll();
      applyStock(readSaved() || DEFAULT_STOCK);
    }})();
  </script>
</body>
</html>"""
    out_path.write_text(html_doc, encoding="utf-8")


def renumber_reports(rows: list[RowReport]) -> list[RowReport]:
    """Gán lại cột # hiển thị 1..n."""
    out: list[RowReport] = []
    for i, r in enumerate(rows, start=1):
        out.append(
            RowReport(
                row=i,
                sku=r.sku,
                name=r.name,
                direct=r.direct,
                fee30=r.fee30,
                fee04=r.fee04,
                shopee_price=r.shopee_price,
                old_price=r.old_price,
                delta=r.delta,
                has_product=r.has_product,
                has_images=r.has_images,
                web_url=r.web_url,
                on_shopee=r.on_shopee,
                shopee_url=r.shopee_url,
                thumb_url=r.thumb_url,
                image_url=r.image_url,
                in_stock=r.in_stock,
                stock_label=r.stock_label,
                note=r.note,
            )
        )
    return out


def write_csv_rows(csv_rows: list[dict], csv_path: Path) -> None:
    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "web_id",
                "sku",
                "ten_shopee",
                "da_dang_shopee",
                "gia_le_max",
                "gia_shopee",
                "shopee_item_id",
                "shopee_url",
                "link_web",
            ],
        )
        w.writeheader()
        w.writerows(csv_rows)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    parser.add_argument(
        "--only-new",
        action="store_true",
        help="Chỉ tạo gói SP chưa đăng Shopee (bỏ qua SKU đã có trên shop theo export Seller Centre)",
    )
    parser.add_argument(
        "--seller-export",
        type=Path,
        help="Export mass_update_basic_info từ Seller Centre (đối chiếu đã đăng)",
    )
    args = parser.parse_args()

    if not args.template.is_file():
        raise SystemExit(f"Không thấy template: {args.template}")

    args.out_dir.mkdir(parents=True, exist_ok=True)
    out_xlsx_full = args.out_dir / "Shopee_mass_upload_71sp_READY.xlsx"
    out_xlsx_new = args.out_dir / "Shopee_mass_upload_CHUA_DANG_READY.xlsx"
    shutil.copy2(args.template, out_xlsx_full)
    out_xlsx = out_xlsx_new if args.only_new else out_xlsx_full

    products = parse_products_js()
    manifest = parse_manifest()
    enrich_images(products, manifest)

    if args.seller_export and args.seller_export.is_file():
        dest = args.out_dir / "seller-doi-chieu-basic-info.xlsx"
        if args.seller_export.resolve() != dest.resolve():
            shutil.copy2(args.seller_export, dest)

    export_links, export_item_ids, doi_chieu_note = load_posted_from_seller_export(args.out_dir)
    shopee_links = load_shopee_links()
    # URL: ưu tiên export shop (đối chiếu), fallback link trên web
    all_links = {**shopee_links, **export_links}

    stock_map = load_stock_status(args.out_dir)
    price_overrides = load_price_overrides(args.out_dir)
    stock_note = (
        "<strong>Còn hàng:</strong> tích ✓ trên báo cáo → bấm <strong>Lưu tồn kho</strong> "
        "(lưu trình duyệt + tải <code>stock-status.json</code>)."
    )

    wb = openpyxl.load_workbook(out_xlsx_full)
    ws = wb[SHEET]
    reports: list[RowReport] = []
    csv_rows: list[dict] = []

    formula_note = (
        "Giá lẻ bán = mức cao nhất trong khoảng (vd. 29–35k → 35.000đ). "
        f"Giá Shopee = lẻ × 1,30 × 1,04 ≈ × {(1 + FEE_PLATFORM) * (1 + FEE_RETURN):.3f}, "
        f"làm tròn lên bội {ROUND_STEP:,}đ.".replace(",", ".")
    )

    for row in range(DATA_START_ROW, ws.max_row + 1):
        sku_val = ws.cell(row=row, column=COL_SKU).value
        if not sku_val:
            continue
        sku = str(sku_val).strip()
        old_name = str(ws.cell(row=row, column=COL_NAME).value or "")
        try:
            old_price = int(float(ws.cell(row=row, column=COL_PRICE).value or 0))
        except (TypeError, ValueError):
            old_price = None

        prod = products.get(sku)
        direct = parse_direct_retail_vnd(prod) if prod else None
        if sku in price_overrides:
            direct = price_overrides[sku]
        shopee_price = fee30 = fee04 = None
        note = ""
        on_shopee = is_on_shopee(sku, export_links, export_item_ids, all_links)
        shopee_url = export_links.get(sku) or all_links.get(sku, "")

        if prod and direct:
            shopee_price, fee30, fee04, _ = resolve_shopee_listing_price(direct)
            ws.cell(row=row, column=COL_NAME).value = build_shopee_title(old_name, prod)
            ws.cell(row=row, column=COL_DESC).value = build_shopee_description(prod, shopee_price)
            ws.cell(row=row, column=COL_PRICE).value = str(shopee_price)

            imgs = prod.images[:8]
            if imgs:
                cover = abs_url(imgs[0])
                ws.cell(row=row, column=COL_COVER).value = cover
                if imgs:
                    ws.cell(row=row, column=COL_VAR_IMG).value = cover
                for i, img in enumerate(imgs):
                    ws.cell(row=row, column=COL_IMG_START + i).value = abs_url(img)
        elif prod:
            note = "Chưa parse được giá lẻ 1–10 — giữ giá Excel cũ"
            shopee_price = old_price
        else:
            note = "Không có trong products.js"

        enable_all_shipping_channels(ws, row)
        desc_val = ws.cell(row=row, column=COL_DESC).value
        if desc_val:
            ws.cell(row=row, column=COL_DESC).value = sanitize_shopee_text(str(desc_val))

        delta = (shopee_price - old_price) if shopee_price is not None and old_price else None
        in_stock, stock_label = resolve_stock_status(sku, stock_map)
        reports.append(
            RowReport(
                row=row,
                sku=sku,
                name=old_name,
                direct=direct,
                fee30=fee30 or 0,
                fee04=fee04 or 0,
                shopee_price=shopee_price,
                old_price=old_price,
                delta=delta,
                has_product=bool(prod),
                has_images=bool(prod and prod.images),
                web_url=product_page_url(sku),
                on_shopee=on_shopee,
                shopee_url=shopee_url,
                thumb_url=product_thumb_url(prod),
                image_url=abs_url(prod.images[0]) if prod and prod.images else "",
                in_stock=in_stock,
                stock_label=stock_label,
                note=note,
            )
        )
        csv_rows.append(
            {
                "web_id": sku,
                "sku": sku,
                "ten_shopee": ws.cell(row=row, column=COL_NAME).value or old_name,
                "da_dang_shopee": "Có" if on_shopee else "Chưa",
                "gia_le_max": direct or "",
                "gia_shopee": shopee_price or old_price or "",
                "shopee_item_id": shopee_url.rstrip("/").split("/")[-1] if on_shopee else "",
                "shopee_url": shopee_url,
                "link_web": product_page_url(sku),
            }
        )

    gia_csv_path = args.out_dir / "gia-le-cap-nhat.csv"
    gia_rows = load_gia_le_csv(gia_csv_path)
    if gia_rows:
        shopee_from_json = load_shopee_prices_json(args.out_dir)
        allowed = {g.sku for g in gia_rows}
        reports = filter_reports_by_gia_le(reports, gia_rows, shopee_from_json)
        csv_rows = [r for r in csv_rows if r["sku"] in allowed]
        csv_by_sku = {r["sku"]: r for r in csv_rows}
        for g in gia_rows:
            row = csv_by_sku.get(g.sku)
            if row:
                row["ten_shopee"] = g.ten
                row["gia_le_max"] = g.gia_le
                sp = shopee_from_json.get(g.sku) or g.gia_shopee
                if sp:
                    row["gia_shopee"] = sp

    posted_excel_rows = sorted((r.row for r in reports if r.on_shopee), reverse=True)
    new_reports = [r for r in reports if not r.on_shopee]

    for row in csv_rows:
        sku = row["sku"]
        if sku in export_links:
            row["shopee_url"] = export_links[sku]
            row["shopee_item_id"] = shopee_item_id(export_links[sku])
            row["da_dang_shopee"] = "Có"
        elif is_on_shopee(sku, export_links, export_item_ids, all_links):
            row["shopee_url"] = all_links[sku]
            row["shopee_item_id"] = shopee_item_id(all_links[sku])
            row["da_dang_shopee"] = "Có"
    posted_src = "export Seller Centre" if export_item_ids else "shopee-links.js"
    skip_posted = (
        f"<strong>Đã bỏ qua {len(posted_excel_rows)} SP</strong> đã có trên shop "
        f"(đối chiếu {posted_src}) — chỉ còn SP cần đăng mới."
    )

    new_csv_rows = [r for r in csv_rows if r["da_dang_shopee"] != "Có"]

    if args.only_new:
        for row_idx in posted_excel_rows:
            ws.delete_rows(row_idx)
        wb.save(out_xlsx_new)
        html_path = args.out_dir / "bao-cao-gia-shopee-chua-dang.html"
        csv_path = args.out_dir / "shopee-item-ids-chua-dang.csv"
        write_html_report(
            renumber_reports(new_reports),
            html_path,
            formula_note,
            excel_name=out_xlsx_new.name,
            page_title="SP chưa đăng Shopee — upload mùa Trung Thu 2026",
            skip_note=skip_posted,
            doi_chieu_note=doi_chieu_note,
            stock_note=stock_note,
        )
        write_csv_rows(new_csv_rows, csv_path)
        summary = {
            "generated": date.today().isoformat(),
            "formula": formula_note,
            "excel": str(out_xlsx_new),
            "html": str(html_path),
            "csv": str(csv_path),
            "rows": len(new_reports),
            "skipped_posted": len(posted_excel_rows),
            "priced": sum(1 for r in new_reports if r.direct),
        }
    else:
        wb.save(out_xlsx_full)
        html_path = args.out_dir / "bao-cao-gia-shopee.html"
        csv_path = args.out_dir / "shopee-item-ids.csv"
        write_html_report(reports, html_path, formula_note, doi_chieu_note=doi_chieu_note, stock_note=stock_note)
        write_gia_le_csv(reports, args.out_dir / "gia-le-cap-nhat.csv")
        write_csv_rows(csv_rows, csv_path)

        # Gói upload: chỉ SP chưa đăng
        wb_new = openpyxl.load_workbook(out_xlsx_full)
        ws_new = wb_new[SHEET]
        for row_idx in posted_excel_rows:
            ws_new.delete_rows(row_idx)
        wb_new.save(out_xlsx_new)
        html_new = args.out_dir / "bao-cao-gia-shopee-chua-dang.html"
        csv_new = args.out_dir / "shopee-item-ids-chua-dang.csv"
        write_html_report(
            renumber_reports(new_reports),
            html_new,
            formula_note,
            excel_name=out_xlsx_new.name,
            page_title="SP chưa đăng Shopee — upload mùa Trung Thu 2026",
            skip_note=skip_posted,
            doi_chieu_note=doi_chieu_note,
            stock_note=stock_note,
        )
        write_csv_rows(new_csv_rows, csv_new)

        summary = {
            "generated": date.today().isoformat(),
            "formula": formula_note,
            "excel_full": str(out_xlsx_full),
            "excel_chua_dang": str(out_xlsx_new),
            "html_full": str(html_path),
            "html_chua_dang": str(html_new),
            "csv_full": str(csv_path),
            "csv_chua_dang": str(csv_new),
            "rows": len(reports),
            "rows_chua_dang": len(new_reports),
            "skipped_posted": len(posted_excel_rows),
            "priced": sum(1 for r in reports if r.direct),
        }

    (args.out_dir / "manifest.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )

    if args.only_new:
        print(f"✅ Excel (chưa đăng) → {out_xlsx_new}")
        print(f"✅ Báo cáo HTML → {html_path}")
        print(f"✅ CSV item IDs → {csv_path}")
        print(f"   Bỏ qua {len(posted_excel_rows)} SP đã đăng · {len(new_reports)} SP sẵn upload")
    else:
        print(f"✅ Excel đầy đủ → {out_xlsx_full}")
        print(f"✅ Excel chưa đăng → {out_xlsx_new} ({len(new_reports)} SP)")
        print(f"✅ Báo cáo HTML → {html_path} + {html_new.name}")
        print(f"✅ CSV item IDs → {csv_path} + {csv_new.name}")
        print(f"   {summary['priced']}/{summary['rows']} có giá · bỏ qua {len(posted_excel_rows)} đã đăng")


if __name__ == "__main__":
    main()
