#!/usr/bin/env python3
"""Đọc export Seller Centre (basic/sales info) → sync link web + tạo basic_info sẵn upload [Mẫu 2026]."""
from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import zipfile
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
PRODUCTS_JS = ROOT / "js" / "products.js"
OUT_JS = ROOT / "js" / "shopee-links.js"
OUT_MAP = ROOT / "data" / "shopee" / "product_mapping.json"
SHOP_ID = 1307955653
PREFIX = "[Mẫu 2026]"

DATA_START_ROW = 7  # 1-based Excel (sau header Shopee)
COL_PID = 1
COL_SKU = 2
COL_NAME = 3


def _hop_qua_root() -> Path:
    here = Path(__file__).resolve()
    for parent in here.parents:
        if (parent / "Shopee_mass_upload_71sp.xlsx").is_file():
            return parent
        if (parent / "hop-qua" / "Shopee_mass_upload_71sp.xlsx").is_file():
            return parent / "hop-qua"
    return Path("/home/vananh/huong-dan/du-an/vo-anh/hop-qua")


HOP_QUA = _hop_qua_root()
OUT_DIR = HOP_QUA / "shopee-upload-2026"


def patch_xlsx_for_openpyxl(src: Path, dst: Path) -> Path:
    """Sửa activePane lỗi openpyxl trong file export Shopee."""
    with zipfile.ZipFile(src, "r") as zin:
        with zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename.startswith("xl/worksheets/") and item.filename.endswith(".xml"):
                    text = data.decode("utf-8")
                    text = re.sub(r'activePane="[^"]*"', 'activePane="topLeft"', text)
                    data = text.encode("utf-8")
                zout.writestr(item, data)
    return dst


def parse_products() -> dict[str, dict]:
    text = PRODUCTS_JS.read_text(encoding="utf-8")
    out: dict[str, dict] = {}
    for block in re.split(r"\n\s*\{", text)[1:]:
        pid = re.search(r"id:\s*'([^']+)'", block)
        if not pid:
            continue
        folder = re.search(r"folder:\s*'([^']*)'", block)
        name = re.search(r"name:\s*'((?:\\'|[^'])*)'", block)
        out[pid.group(1)] = {
            "folder": (folder.group(1) if folder else ""),
            "name": (name.group(1) if name else "").replace("\\'", "'"),
        }
    return out


def is_item_id(val: object) -> bool:
    s = str(val or "").strip()
    return bool(re.fullmatch(r"\d{8,12}", s))


def item_url(item_id: str) -> str:
    return f"https://shopee.vn/product/{SHOP_ID}/{item_id.strip()}"


def build_listing_title(sku: str, current: str, catalog: dict[str, dict]) -> str | None:
    """Trả về tên mới hoặc None nếu giữ nguyên."""
    cur = str(current or "").strip()
    if not cur:
        return None
    if PREFIX.lower() in cur.lower() or cur.lower().startswith("[mẫu"):
        return None
    prod = catalog.get(sku)
    if not prod:
        return None
    if prod["folder"].startswith("cap-nhat-2026/"):
        short = prod["name"][:90]
        title = f"{PREFIX} Vỏ hộp trung thu {short}"
    else:
        title = f"{PREFIX} {prod['name'] or cur}"
    return title[:120]


def read_basic_links(basic_path: Path) -> dict[str, str]:
    """SKU → URL từ export basic info."""
    df = pd.read_excel(basic_path, engine="calamine", header=None)
    links: dict[str, str] = {}
    for i in range(len(df)):
        pid = df.iloc[i, 0]
        if not is_item_id(pid):
            continue
        sku = str(df.iloc[i, 1] or "").strip()
        if not sku or sku in ("nan", "SKU Sản phẩm", "et_title_parent_sku"):
            continue
        if re.fullmatch(r"[a-f0-9]{32}", sku):
            continue
        links[sku] = item_url(str(int(float(pid))))
    return links


def read_existing_js() -> dict[str, str]:
    if not OUT_JS.is_file():
        return {}
    return dict(re.findall(r"'([^']+)':\s*'(https://shopee\.vn/product/\d+/\d+)'", OUT_JS.read_text(encoding="utf-8")))


def write_js(urls: dict[str, str]) -> None:
    lines = [
        "// Link Shopee trực tiếp theo sản phẩm (shop longthibo958)",
        "// Cập nhật: python3 scripts/prepare-shopee-mass-update-export.py",
        "const SHOPEE_PRODUCT_URLS = {",
    ]
    for pid in sorted(urls):
        lines.append(f"    '{pid}': '{urls[pid]}',")
    lines.append("};")
    lines.append("")
    OUT_JS.write_text("\n".join(lines), encoding="utf-8")


def update_csv_files(urls: dict[str, str]) -> None:
    for csv_path in (OUT_DIR / "shopee-item-ids.csv", OUT_DIR / "shopee-item-ids-chua-dang.csv"):
        if not csv_path.is_file():
            continue
        rows: list[dict] = []
        with csv_path.open(encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            fieldnames = list(reader.fieldnames or [])
            for row in reader:
                sku = (row.get("sku") or row.get("web_id") or "").strip()
                if sku in urls:
                    row["shopee_url"] = urls[sku]
                    row["shopee_item_id"] = urls[sku].rstrip("/").split("/")[-1]
                    row["da_dang_shopee"] = "Có"
                rows.append(row)
        with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader()
            w.writerows(rows)


def prepare_basic_info(basic_src: Path, out_path: Path, catalog: dict[str, dict]) -> tuple[int, int]:
    patched = out_path.with_suffix(".patched.xlsx")
    patch_xlsx_for_openpyxl(basic_src, patched)
    wb = openpyxl.load_workbook(patched)
    ws = wb.active
    renamed = 0
    skipped = 0
    for row in range(DATA_START_ROW, ws.max_row + 1):
        pid = ws.cell(row=row, column=COL_PID).value
        if not is_item_id(pid):
            continue
        sku = str(ws.cell(row=row, column=COL_SKU).value or "").strip()
        if not sku or sku in ("SKU Sản phẩm",) or re.fullmatch(r"[a-f0-9]{32}", sku):
            skipped += 1
            continue
        old_name = str(ws.cell(row=row, column=COL_NAME).value or "")
        new_name = build_listing_title(sku, old_name, catalog)
        if new_name:
            ws.cell(row=row, column=COL_NAME).value = new_name
            renamed += 1
    wb.save(out_path)
    patched.unlink(missing_ok=True)
    return renamed, skipped


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--basic", type=Path, required=True, help="mass_update_basic_info_*.xlsx")
    parser.add_argument("--sales", type=Path, help="mass_update_sales_info_*.xlsx (tham khảo)")
    parser.add_argument("--out-dir", type=Path, default=OUT_DIR)
    args = parser.parse_args()

    if not args.basic.is_file():
        raise SystemExit(f"Không thấy: {args.basic}")

    args.out_dir.mkdir(parents=True, exist_ok=True)
    catalog = parse_products()

    links = read_basic_links(args.basic)
    urls = read_existing_js()
    for sku, url in links.items():
        urls[sku] = url

    write_js(urls)
    mappings = [
        {"productId": k, "shopeeUrl": v, "itemid": int(v.rstrip("/").split("/")[-1])}
        for k, v in sorted(urls.items())
    ]
    OUT_MAP.parent.mkdir(parents=True, exist_ok=True)
    OUT_MAP.write_text(
        json.dumps({"mapped": mappings, "source": str(args.basic)}, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    update_csv_files(urls)

    out_basic = args.out_dir / "mass_update_basic_info_READY.xlsx"
    renamed, skipped = prepare_basic_info(args.basic, out_basic, catalog)

    if args.sales and args.sales.is_file():
        shutil.copy2(args.sales, args.out_dir / args.sales.name)

    shutil.copy2(args.basic, args.out_dir / "mass_update_basic_info_SOURCE.xlsx")

    print(f"✅ Link Shopee: {len(urls)} SP → {OUT_JS}")
    print(f"✅ Basic info READY: {out_basic}")
    print(f"   Đổi tên thêm {PREFIX}: {renamed} dòng · giữ nguyên: {skipped + (len(links) - renamed)}")
    print("   Upload file READY lên Seller Centre → Cập nhật hàng loạt → tab Tải lên")


if __name__ == "__main__":
    main()
