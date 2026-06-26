#!/usr/bin/env python3
"""Đồng bộ link Shopee lên website sau khi mass upload.

Nguồn (theo thứ tự ưu tiên):
1. File export Seller Centre «Sales Info» / «Thông tin bán hàng» (.xlsx)
2. seller-links.tsv — mỗi dòng: sku<TAB>item_id hoặc sku<TAB>url
3. shopee-item-ids.csv / shopee-item-ids-chua-dang.csv (cột shopee_item_id)

Chạy:
  python3 scripts/sync-shopee-links-after-upload.py
  python3 scripts/sync-shopee-links-after-upload.py --export ~/Downloads/sales_info.xlsx
"""
from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
OUT_JS = ROOT / "js" / "shopee-links.js"
OUT_MAP = ROOT / "data" / "shopee" / "product_mapping.json"
SHOP_ID = 1307955653


def _hop_qua_root() -> Path:
    here = Path(__file__).resolve()
    for parent in here.parents:
        if (parent / "Shopee_mass_upload_71sp.xlsx").is_file():
            return parent
        if (parent / "hop-qua" / "Shopee_mass_upload_71sp.xlsx").is_file():
            return parent / "hop-qua"
    return Path("/home/vananh/huong-dan/du-an/vo-anh/hop-qua")


HOP_QUA = _hop_qua_root()
UPLOAD_DIR = HOP_QUA / "shopee-upload-2026"
DEFAULT_EXPORT_NAMES = (
    "seller-sales-info.xlsx",
    "Sales_Info.xlsx",
    "sales_info.xlsx",
    "Thong_tin_ban_hang.xlsx",
)
DEFAULT_TSV = UPLOAD_DIR / "seller-links.tsv"
DEFAULT_CSV = UPLOAD_DIR / "shopee-item-ids.csv"
DEFAULT_CSV_NEW = UPLOAD_DIR / "shopee-item-ids-chua-dang.csv"

PID_HEADERS = {
    "product id",
    "id sản phẩm",
    "mã sản phẩm shopee",
    "item id",
    "id sản phẩm",
}
SKU_HEADERS = {
    "parent sku",
    "parent sku #",
    "sku sản phẩm",
    "sku sản phẩm cha",
    "sku cha",
    "parent_sku",
}
NAME_HEADERS = {"product name", "tên sản phẩm", "ten san pham"}


def norm_header(v: object) -> str:
    return re.sub(r"\s+", " ", str(v or "").strip().lower())


def is_shopee_item_id(val: object) -> bool:
    s = str(val or "").strip()
    return bool(re.fullmatch(r"\d{8,12}", s))


def item_url(item_id: str) -> str:
    return f"https://shopee.vn/product/{SHOP_ID}/{item_id.strip()}"


def parse_item_id(raw: str) -> str:
    raw = raw.strip()
    m = re.search(r"/product/\d+/(\d+)", raw)
    if m:
        return m.group(1)
    if is_shopee_item_id(raw):
        return raw
    return ""


def read_existing_js() -> dict[str, str]:
    if not OUT_JS.is_file():
        return {}
    return dict(re.findall(r"'([^']+)':\s*'(https://shopee\.vn/product/\d+/\d+)'", OUT_JS.read_text(encoding="utf-8")))


def load_csv_links(path: Path) -> dict[str, str]:
    if not path.is_file():
        return {}
    out: dict[str, str] = {}
    with path.open(encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            web_id = (row.get("web_id") or row.get("sku") or "").strip()
            if not web_id:
                continue
            item_id = (row.get("shopee_item_id") or "").strip()
            url = (row.get("shopee_url") or "").strip()
            if item_id and not url:
                url = item_url(item_id)
            if url and "shopee.vn/product/" in url:
                out[web_id] = url
    return out


def load_tsv_links(path: Path) -> dict[str, str]:
    if not path.is_file():
        return {}
    out: dict[str, str] = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        parts = re.split(r"\t+", line) if "\t" in line else line.split(",", 1)
        if len(parts) < 2:
            continue
        sku, val = parts[0].strip(), parts[1].strip()
        iid = parse_item_id(val)
        if sku and iid:
            out[sku] = item_url(iid)
    return out


def find_export_paths(explicit: Path | None) -> list[Path]:
    if explicit and explicit.is_file():
        return [explicit]
    found: list[Path] = []
    for name in DEFAULT_EXPORT_NAMES:
        p = UPLOAD_DIR / name
        if p.is_file():
            found.append(p)
    for base in (Path.home() / "Downloads", Path.home() / "Downloads" / "Shopee", UPLOAD_DIR):
        if not base.is_dir():
            continue
        for p in sorted(base.glob("*.xlsx"), key=lambda x: x.stat().st_mtime, reverse=True):
            if p.name.startswith("Shopee_mass_upload"):
                continue
            found.append(p)
    # unique preserve order
    seen: set[Path] = set()
    out: list[Path] = []
    for p in found:
        if p in seen:
            continue
        seen.add(p)
        out.append(p)
    return out


def parse_seller_export(path: Path) -> dict[str, str]:
    """Đọc export Sales Info — ghép Parent SKU → Product ID."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    mappings: dict[str, str] = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_row < 2:
            continue
        hdr_row = None
        col_pid = col_sku = 0
        for r in range(1, min(15, max_row + 1)):
            for c in range(1, min(max_col + 1, 60)):
                h = norm_header(ws.cell(r, c).value)
                if h in PID_HEADERS or h == "product id":
                    col_pid = c
                    hdr_row = r
                if h in SKU_HEADERS:
                    col_sku = c
                    hdr_row = hdr_row or r
            if col_pid and col_sku:
                break
        if not col_pid or not hdr_row:
            continue
        for r in range(hdr_row + 1, max_row + 1):
            pid = ws.cell(r, col_pid).value
            if not is_shopee_item_id(pid):
                continue
            sku = ""
            if col_sku:
                sku = str(ws.cell(r, col_sku).value or "").strip()
            if not sku:
                continue
            mappings[sku] = item_url(str(int(pid)))
    wb.close()
    return mappings


def write_js(urls: dict[str, str]) -> None:
    lines = [
        "// Link Shopee trực tiếp theo sản phẩm (shop longthibo958)",
        "// Cập nhật: python3 scripts/sync-shopee-links-after-upload.py",
        "const SHOPEE_PRODUCT_URLS = {",
    ]
    for pid in sorted(urls):
        lines.append(f"    '{pid}': '{urls[pid]}',")
    lines.append("};")
    lines.append("")
    OUT_JS.write_text("\n".join(lines), encoding="utf-8")


def update_csv_files(urls: dict[str, str]) -> int:
    updated = 0
    for csv_path in (DEFAULT_CSV, DEFAULT_CSV_NEW):
        if not csv_path.is_file():
            continue
        rows: list[dict] = []
        with csv_path.open(encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            fieldnames = reader.fieldnames or []
            for row in reader:
                sku = (row.get("sku") or row.get("web_id") or "").strip()
                if sku in urls:
                    iid = urls[sku].rstrip("/").split("/")[-1]
                    row["shopee_url"] = urls[sku]
                    row["shopee_item_id"] = iid
                    row["da_dang_shopee"] = "Có"
                    updated += 1
                rows.append(row)
        if rows and fieldnames:
            with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
                w = csv.DictWriter(f, fieldnames=fieldnames)
                w.writeheader()
                w.writerows(rows)
    return updated


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--export", type=Path, help="File Sales Info export từ Seller Centre")
    parser.add_argument("--tsv", type=Path, default=DEFAULT_TSV)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    urls = read_existing_js()
    sources: list[tuple[str, dict[str, str]]] = []

    for path in find_export_paths(args.export):
        parsed = parse_seller_export(path)
        if parsed:
            sources.append((f"export:{path.name}", parsed))

    tsv_links = load_tsv_links(args.tsv)
    if tsv_links:
        sources.append((f"tsv:{args.tsv.name}", tsv_links))

    for csv_path in (DEFAULT_CSV, DEFAULT_CSV_NEW):
        csv_links = load_csv_links(csv_path)
        if csv_links:
            sources.append((f"csv:{csv_path.name}", csv_links))

    added = 0
    for label, batch in sources:
        for sku, url in batch.items():
            if urls.get(sku) != url:
                urls[sku] = url
                added += 1
        print(f"  + {label}: {len(batch)} link")

    if not added and len(urls) == len(read_existing_js()):
        print("⚠️  Chưa có link mới.")
        print("   Export Seller Centre → Sản phẩm → Công cụ hàng loạt → Cập nhật hàng loạt")
        print("   → tab Tải xuống → «Thông tin bán hàng» / Sales Info → Tạo → Tải về")
        print(f"   → lưu vào: {UPLOAD_DIR / 'seller-sales-info.xlsx'}")
        print(f"   Hoặc điền: {DEFAULT_TSV}  (sku<TAB>item_id)")
        print("   Rồi chạy lại script này.")
        return

    mappings = []
    for sku, url in sorted(urls.items()):
        mappings.append(
            {
                "productId": sku,
                "shopeeUrl": url,
                "itemid": int(url.rstrip("/").split("/")[-1]),
            }
        )

    if args.dry_run:
        print(f"[dry-run] {len(urls)} link tổng (+{added} mới)")
        return

    write_js(urls)
    OUT_MAP.parent.mkdir(parents=True, exist_ok=True)
    OUT_MAP.write_text(
        json.dumps({"mapped": mappings, "sources": [s[0] for s in sources]}, ensure_ascii=False, indent=2)
        + "\n",
        encoding="utf-8",
    )
    csv_n = update_csv_files(urls)
    print(f"✅ {len(urls)} link → {OUT_JS} (+{added} mới)")
    print(f"✅ Cập nhật {csv_n} dòng CSV")
    print("   Tiếp theo: git add js/shopee-links.js data/shopee/product_mapping.json && git push")


if __name__ == "__main__":
    main()
